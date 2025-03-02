import json
import imaplib
import email
import os
import requests
import time
import datetime
import pandas as pd  # Import pandas for Excel export
from email.header import decode_header
from dotenv import load_dotenv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

load_dotenv()

# Your email credentials
EMAIL = os.getenv("EMAIL")
PASSWORD = os.getenv("PASSWORD")
IMAP_SERVER = "imap.gmail.com"  # Example: 'imap.gmail.com' for Gmail

# PDF.co API key
API_KEY = os.getenv("API_KEY")

# Base URL for PDF.co Web API requests
BASE_URL = "https://api.pdf.co/v1"


def connect_email():
    """Connect to the email server and return the mail object."""
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, PASSWORD)
    return mail


def search_invoices(mail):
    """Search for emails with the subject 'invoice'."""
    mail.select("inbox")
    status, messages = mail.search(None, '(SUBJECT "invoice")')
    email_ids = messages[0].split()
    return email_ids


def download_pdf_attachment(mail, email_id):
    """Download PDF attachment from the email."""
    status, data = mail.fetch(email_id, "(RFC822)")
    for response_part in data:
        if isinstance(response_part, tuple):
            msg = email.message_from_bytes(response_part[1])
            for part in msg.walk():
                if part.get_content_type() == "application/pdf":
                    filename = part.get_filename()
                    if filename:
                        filename = decode_header(filename)[0][0]
                        if isinstance(filename, bytes):
                            filename = filename.decode()
                        with open(filename, "wb") as f:
                            f.write(part.get_payload(decode=True))
                            return filename
    return None


def get_parsed_invoice(file_path):
    """Send the invoice PDF to PDF.co for processing and return the parsed data."""
    with open(file_path, "rb") as file:
        # Upload the file to PDF.co and get the file URL
        url = f"{BASE_URL}/file/upload/get-presigned-url?name={os.path.basename(file_path)}"
        headers = {"x-api-key": API_KEY}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            json_response = response.json()
            if json_response["error"] == False:
                upload_url = json_response["presignedUrl"]
                file_url = json_response["url"]
                
                # Upload the actual file to the presigned URL
                with open(file_path, "rb") as file:
                    upload_response = requests.put(upload_url, data=file)
                    if upload_response.status_code == 200:
                        return parse_invoice(file_url)
            else:
                print("Error uploading file:", json_response["message"])
    return None


def parse_invoice(uploaded_file_url):
    """Send the invoice to PDF.co AI Invoice Parser API."""
    parameters = {"url": uploaded_file_url}
    url = f"{BASE_URL}/ai-invoice-parser"
    headers = {"x-api-key": API_KEY}
    response = requests.post(url, data=parameters, headers=headers)
    if response.status_code == 200:
        json_response = response.json()
        if json_response["error"] == False:
            job_id = json_response["jobId"]
            return check_job_status(job_id)
        else:
            print("Error in parsing invoice:", json_response["message"])
    else:
        print("Error with request:", response.status_code, response.reason)
    return None


def check_job_status(job_id):
    """Check the job status until the job is finished."""
    while True:
        url = f"{BASE_URL}/job/check?jobid={job_id}"
        headers = {"x-api-key": API_KEY}
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            json_response = response.json()
            status = json_response["status"]
            print(f"Job status: {status}")
            if status == "success":
                # The job has finished, return the result
                return json_response
            elif status == "working":
                time.sleep(3)  # Wait a bit before checking again
            else:
                print(f"Job failed: {status}")
                break
        else:
            print("Error with job status request:", response.status_code, response.reason)
            break
    return None


def flatten_data(data, parent_key='', sep='_'):
    """Flatten nested dictionaries into a flat dictionary."""
    items = []
    for key, value in data.items():
        new_key = f"{parent_key}{sep}{key}" if parent_key else key
        if isinstance(value, dict):
            items.extend(flatten_data(value, new_key, sep=sep).items())
        else:
            items.append((new_key, value))
    return dict(items)


def save_json_and_excel(data, file_name):
    """Save parsed data to both a JSON file and an Excel file with formatting."""
    
    # Save data to JSON
    with open(file_name + ".json", "w") as json_file:
        json.dump(data, json_file, indent=4)
    print(f"Saved parsed invoice to {file_name}.json")

    # Flatten the data if it's nested
    flat_data = flatten_data(data)

    # Convert the flattened data into a pandas DataFrame
    try:
        # Convert the data into a DataFrame (this assumes the data is a dictionary)
        df = pd.DataFrame([flat_data])  # wrap in list to ensure it's treated as a single row

        # Save the DataFrame to Excel using a context manager
        excel_filename = file_name + ".xlsx"
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Parsed Invoice")
            
            # Get the worksheet
            worksheet = writer.sheets["Parsed Invoice"]
            
            # Apply formatting and styling
            # Set header bold and center-aligned
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal="center")

            # Auto-adjust column widths and set some padding
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 5)  # Add padding
                worksheet.column_dimensions[column].width = adjusted_width

            # Apply borders around the cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(border_style="thin"),
                        right=openpyxl.styles.Side(border_style="thin"),
                        top=openpyxl.styles.Side(border_style="thin"),
                        bottom=openpyxl.styles.Side(border_style="thin")
                    )

            # Apply alternating row colors (for better readability)
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                if idx % 2 == 0:
                    for cell in row:
                        cell.fill = openpyxl.styles.PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

        print(f"Saved parsed invoice to {excel_filename}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")


def main():
    # Step 1: Connect to email
    mail = connect_email()

    # Step 2: Search for invoice emails
    email_ids = search_invoices(mail)
    if not email_ids:
        print("No invoices found.")
        return

    # Step 3: Process each invoice email
    for email_id in email_ids:
        # Step 4: Download the PDF attachment from the email
        pdf_filename = download_pdf_attachment(mail, email_id)
        if pdf_filename:
            print(f"Downloaded PDF: {pdf_filename}")

            # Step 5: Use PDF.co API to process the invoice
            parsed_data = get_parsed_invoice(pdf_filename)
            if parsed_data:
                file_name = pdf_filename.replace(".pdf", "")
                # Step 6: Save the parsed invoice data to both JSON and Excel files
                save_json_and_excel(parsed_data, file_name)
            else:
                print(f"Failed to parse invoice: {pdf_filename}")
        else:
            print(f"No PDF found in email {email_id}")


if __name__ == "__main__":
    main()









