import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Your JSON data
data = {
    "status": "success",
    "message": "Success",
    "pageCount": 2,
    "url": "https://pdf-temp-files.s3.us-west-2.amazonaws.com/YL6AMKWOS3L8896OI9F5VHHFKVGVDBQO--155-200/invoice.json?X-Amz-Expires=3600&X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Credential=AKIAIZJDPLX6D7EHVCKA/20250302/us-west-2/s3/aws4_request&X-Amz-Date=20250302T053755Z&X-Amz-SignedHeaders=host&X-Amz-Signature=63c6c474ec5bed87d2d49171c6f889ed2f5bb8108d2442bfdbf50bbf10359940",
    "body": {
        "vendor": {
            "name": "",
            "address": {
                "streetAddress": "",
                "city": "",
                "state": "",
                "postalCode": "",
                "country": ""
            },
            "contactInformation": {
                "phone": "",
                "fax": "",
                "email": ""
            },
            "entityId": ""
        },
        "customer": {
            "billTo": {
                "name": "David Ramirez",
                "address": {
                    "streetAddress": "65 Oakland Ave",
                    "city": "West Hempstead",
                    "state": "NY",
                    "postalCode": "11552",
                    "country": ""
                },
                "contactInformation": {
                    "phone": "",
                    "email": ""
                },
                "entityId": ""
            },
            "shipTo": {
                "name": "David Ramirez",
                "address": {
                    "streetAddress": "65 Oakland Ave",
                    "city": "West Hempstead",
                    "state": "NY",
                    "postalCode": "11552",
                    "country": ""
                },
                "contactInformation": {
                    "phone": "",
                    "email": ""
                }
            }
        },
        "invoice": {
            "invoiceNo": "19810500",
            "invoiceDate": "26-SEP-2022",
            "poNo": "David Ramirez26-SEP-22 17:52:40",
            "orderNo": "49174389",
            "deliveryDate": ""
        },
        "paymentDetails": {
            "paymentTerms": "IMMEDIATE",
            "dueDate": "26-SEP-2022",
            "total": "USD $81.46",
            "subtotal": "74.99",
            "tax": "6.47",
            "discount": "",
            "shipping": ""
        },
        "lineItems": [
            {
                "order_number": "49174389",
                "po_number": "David Ramirez26-SEP-22 17:52:40",
                "item_number": "9780135813379",
                "item_description": "MyLab Math with Pearson eText -- Instant Access -- for Precalculus Enhanced with Graphing Utilities (18-Weeks)",
                "returnable": "No",
                "quantity": "1",
                "list_price": "74.99",
                "net_price": "74.99",
                "tax": "6.47",
                "line_total": "81.46"
            }
        ]
    }
}

# Create a new workbook
wb = openpyxl.Workbook()

# Function to safely retrieve nested data with default value if key is missing
def safe_get(data, keys, default_value="N/A"):
    for key in keys:
        if isinstance(data, dict) and key in data:
            data = data[key]
        else:
            return default_value
    return data

# Function to add header row with bold font
def add_header_row(ws, row, data):
    for col, value in enumerate(data, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="thin"))

# Create separate sheets for each section

# Vendor Information Sheet
ws_vendor = wb.create_sheet(title="Vendor Info")
add_header_row(ws_vendor, 1, ["Name", "Street Address", "City", "State", "Postal Code", "Country", "Phone", "Fax", "Email"])
ws_vendor.append([safe_get(data, ['body', 'vendor', 'name']),
           safe_get(data, ['body', 'vendor', 'address', 'streetAddress']),
           safe_get(data, ['body', 'vendor', 'address', 'city']),
           safe_get(data, ['body', 'vendor', 'address', 'state']),
           safe_get(data, ['body', 'vendor', 'address', 'postalCode']),
           safe_get(data, ['body', 'vendor', 'address', 'country']),
           safe_get(data, ['body', 'vendor', 'contactInformation', 'phone']),
           safe_get(data, ['body', 'vendor', 'contactInformation', 'fax']),
           safe_get(data, ['body', 'vendor', 'contactInformation', 'email'])])

# Customer Billing Information Sheet
ws_customer_billing = wb.create_sheet(title="Customer Billing")
add_header_row(ws_customer_billing, 1, ["Name", "Street Address", "City", "State", "Postal Code", "Phone", "Email"])
ws_customer_billing.append([safe_get(data, ['body', 'customer', 'billTo', 'name']),
           safe_get(data, ['body', 'customer', 'billTo', 'address', 'streetAddress']),
           safe_get(data, ['body', 'customer', 'billTo', 'address', 'city']),
           safe_get(data, ['body', 'customer', 'billTo', 'address', 'state']),
           safe_get(data, ['body', 'customer', 'billTo', 'address', 'postalCode']),
           safe_get(data, ['body', 'customer', 'billTo', 'contactInformation', 'phone']),
           safe_get(data, ['body', 'customer', 'billTo', 'contactInformation', 'email'])])

# Invoice Details Sheet
ws_invoice = wb.create_sheet(title="Invoice Details")
add_header_row(ws_invoice, 1, ["Invoice No", "Invoice Date", "Order No", "Delivery Date"])
ws_invoice.append([safe_get(data, ['body', 'invoice', 'invoiceNo']),
           safe_get(data, ['body', 'invoice', 'invoiceDate']),
           safe_get(data, ['body', 'invoice', 'orderNo']),
           safe_get(data, ['body', 'invoice', 'deliveryDate'])])

# Line Items Sheet
ws_line_items = wb.create_sheet(title="Line Items")
line_item_headers = ["Order Number", "PO Number", "Item Number", "Item Description", "Returnable", "Quantity", "List Price", "Net Price", "Tax", "Line Total"]
add_header_row(ws_line_items, 1, line_item_headers)

for item in data['body']['lineItems']:
    ws_line_items.append([item['order_number'], item['po_number'], item['item_number'], item['item_description'], item['returnable'], item['quantity'], item['list_price'], item['net_price'], item['tax'], item['line_total']])

# Payment Details Sheet
ws_payment = wb.create_sheet(title="Payment Details")
add_header_row(ws_payment, 1, ["Payment Terms", "Due Date", "Total", "Subtotal", "Tax", "Discount", "Shipping"])
ws_payment.append([safe_get(data, ['body', 'paymentDetails', 'paymentTerms']),
           safe_get(data, ['body', 'paymentDetails', 'dueDate']),
           safe_get(data, ['body', 'paymentDetails', 'total']),
           safe_get(data, ['body', 'paymentDetails', 'subtotal']),
           safe_get(data, ['body', 'paymentDetails', 'tax']),
           safe_get(data, ['body', 'paymentDetails', 'discount']),
           safe_get(data, ['body', 'paymentDetails', 'shipping'])])

# Remove the default sheet created initially
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Save the workbook to a file
wb.save("invoice_data_separate_sheets.xlsx")




